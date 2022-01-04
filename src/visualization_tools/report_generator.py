import os
import re
import shutil
import openpyxl
import pandas

from typing import List


class ReportGenerator:
    def __init__(self):
        self.eft = ExcelFormattingTool()

    def convert_wb(self, old_wb_path, template_wb_path, output_wb_path):
        # Open existing workbook with data
        old_wb = self.eft.load_workbook(old_wb_path)
        # Create copy of template for data
        shutil.copy(template_wb_path, output_wb_path)
        # Copy data from existing workbook into the new template wb
        output_wb = self.eft.load_workbook(output_wb_path)


class ExcelFormattingTool:

    def __init__(self):
        pass

    def load_workbook(self, path: str) -> openpyxl.workbook.Workbook:
        return openpyxl.load_workbook(path)
    
    def append_values(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                      values):
        """
        Adds values to the bottom of the provided sheet
        :param ws:
        :param values:
        :return:
        """
        ws.append(values)

    def create_sheet(self, wb: openpyxl.workbook.Workbook, title, ix):
        wb.create_sheet(title, ix)

    def modify_cell(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                    m, n, value):
        """
        https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html?highlight=iter_cols#openpyxl.worksheet.worksheet.Worksheet.iter_cols
        1-based row and column indexing
        :param sheet:
        :param m:
        :param n:
        :param value:
        :return:
        """
        # 1-based indexing
        sheet.cell(row=m, column=n).value = value

    def get_coord_from_str(self, ix_str):
        """
        https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
        https://stackoverflow.com/questions/12902621/getting-the-row-and-column-numbers-from-coordinate-value-in-openpyxl
        :param ix_str:
        :return:
        """
        row, col = openpyxl.utils.cell.coordinate_to_tuple(ix_str)
        return col, row

    def get_sheet_dims(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        coord_1, coord_2 = re.split(':', ws.calculate_dimension())
        col_1, row_1 = self.get_coord_from_str(coord_1)
        col_2, row_2 = self.get_coord_from_str(coord_2)
        return [col_1, row_1], [col_2, row_2]

    def get_wb_sheet_names(self, wb: openpyxl.workbook.Workbook) -> List[str]:
        return wb.sheetnames

    def replace_col(self, ws, new_col_values):
        pass

    def iter_wb_sheets(self, wb, fxn, **kwargs):
        for sheet in wb:
            fxn(sheet)

    def copy_sheet_in_wb(self, wb: openpyxl.workbook.Workbook,
                         sheet: openpyxl.worksheet.worksheet.Worksheet
                         ) -> openpyxl.worksheet.worksheet.Worksheet:
        # You also cannot copy worksheets between workbooks
        sheet_copy = wb.copy_worksheet(sheet)
        return sheet_copy

    def get_col_length(self):
        pass

def main():
    rg = ReportGenerator()
    eft = ExcelFormattingTool()
    # Paths and filenames
    path = r'C:\Users\gsass\Documents\My Tableau Repository\Datasources\fall_risk_assessments\cohort_1_fall_risk_metrics.xlsx'
    folder = r'C:\Users\gsass\Documents\My Tableau Repository\Datasources\fall_risk_assessments'
    new_file_path = r'C:\Users\gsass\Documents\My Tableau Repository\Datasources\fall_risk_assessments\NEW_cohort_1_fall_risk_metrics.xlsx'

    # Create copy of the spreadsheet
    shutil.copy(path, new_file_path)
    # Open the new copy and modify the sheet values
    wb = openpyxl.load_workbook(new_file_path)
    user_uuids = ['data_abcd-1234']
    # Index the sheet names with UUIDs
    ws = wb[user_uuids[0]]
    ws.cell(row=2, column=11).value = 15
    dim1, dim2 = rg.get_sheet_dims(ws)
    for i in ws.columns:
        print(i)
    # start creating basic functions to copy columns, modify values etc to prep
    # for the conversion of MR stuff

    # data = pandas.read_excel(new_file_path, engine='openpyxl')
    # with open(new_file_path, 'w') as xl_file:
    #     data = pandas.read_excel(xl_file)
    # Save off the modifications made to the document
    wb.save(new_file_path)
    # Publish the dashboard as a PDF


if __name__ == '__main__':
    main()